#!/usr/bin/env python3
"""
Generate TKT Measurement Forms (PDF) from XLSX source data.
Output: formulir/Borang TKT Software.pdf
        formulir/Borang TKT Umum dan Engineering.pdf
"""

import sys, os
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether, PageBreak,
)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FORMULIR_DIR = os.path.join(ROOT, 'formulir')

# Colors
HEADER_BG    = colors.HexColor('#1F3864')
HEADER_FG    = colors.white
SUBHEADER_BG = colors.HexColor('#D6E4F0')
ALT_ROW_BG   = colors.HexColor('#F2F7FB')
BORDER_CLR   = colors.HexColor('#7F8C8D')
RED          = colors.HexColor('#C0392B')
GREEN        = colors.HexColor('#27AE60')
LIGHT_GRAY   = colors.HexColor('#ECF0F1')

# Built-in Helvetica
FN = 'Helvetica'
FB = 'Helvetica-Bold'

def P(text, style):
    return Paragraph(str(text), style)

def make_styles():
    def s(name, **kw):
        kw.setdefault('fontName', FN)
        return ParagraphStyle(name, **kw)
    return {
        'title':    s('title',    fontSize=14, alignment=TA_CENTER, spaceAfter=3),
        'subtitle': s('subtitle', fontSize=9,  alignment=TA_CENTER, spaceAfter=2),
        'section':  s('section',  fontSize=11, spaceAfter=6, textColor=HEADER_BG),
        'label':    s('label',    fontSize=10, leading=13),
        'cell':     s('cell',     fontSize=8,  leading=10),
        'cell_b':   s('cell_b',   fontSize=8,  leading=10, fontName=FB),
        'cell_c':   s('cell_c',   fontSize=8,  leading=10, alignment=TA_CENTER),
        'hdr':      s('hdr',      fontSize=8,  leading=10, fontName=FB, textColor=HEADER_FG),
        'hdr_c':    s('hdr_c',    fontSize=8,  leading=10, fontName=FB, textColor=HEADER_FG, alignment=TA_CENTER),
        'small':    s('small',    fontSize=8,   leading=10),
        'note':     s('note',     fontSize=9,   leading=11),
    }

def read_tkt_data(xlsx_path):
    """Parse TKT levels from XLSX."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    levels = []
    current = None

    for row in range(1, ws.max_row + 1):
        c2 = ws.cell(row=row, column=2).value or ''
        c3 = ws.cell(row=row, column=3).value
        c4 = ws.cell(row=row, column=4).value or ''
        c5 = ws.cell(row=row, column=5).value
        c7 = ws.cell(row=row, column=7).value or ''

        c2_str = str(c2).strip()
        if c2_str.startswith('TKT'):
            try:
                lvl = int(c2_str.split()[-1])
            except ValueError:
                continue
            current = {'level': lvl, 'indicators': [], 'stop_text': ''}
            levels.append(current)
            continue

        if current is None:
            continue

        if isinstance(c3, (int, float)) and c3 != 0:
            ind = {'id': int(c3), 'text': str(c4).strip(), 'keterangan': str(c7).strip()}
            if c5 is not None and c5 != '':
                ind['value'] = c5
            current['indicators'].append(ind)
            if c7.strip():
                current['stop_text'] = str(c7).strip()

        if str(c3).strip() == 'Nilai Rata-rata':
            if c5 is not None and c5 != '':
                current['average'] = float(c5)

    return levels

def build_pdf(xlsx_path, pdf_path, title, subtitle, is_engineering=False):
    """Generate one TKT PDF form."""
    st = make_styles()
    data = read_tkt_data(xlsx_path)

    doc = SimpleDocTemplate(
        pdf_path, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=18*mm,  bottomMargin=15*mm,
    )
    story = []
    W = A4[0] - 36*mm

    # ── Cover page ──────────────────────────────────────
    story.append(P(title, st['title']))
    story.append(P(subtitle, st['subtitle']))
    story.append(Spacer(1, 3*mm))
    story.append(P('KEMENTERIAN RISET, TEKNOLOGI DAN PENDIDIKAN TINGGI', st['subtitle']))
    story.append(P('DIREKTORAT JENDERAL PENGUATAN RISET DAN PENGEMBANGAN', st['subtitle']))
    story.append(P('PROGRAM STUDI TEKNIK INFORMATIKA - INSTITUT TEKNOLOGI SUMATERA', st['subtitle']))
    story.append(Spacer(1, 6*mm))

    profile_fields = [
        'Nama Mahasiswa / Peneliti', 'NIM', 'Program Studi', 'Dosen Pembimbing',
        'Nama Teknologi / Proyek', 'Bidang Teknologi', 'Deskripsi Teknologi', 'Status Riset',
    ]
    profile_data = []
    for f in profile_fields:
        profile_data.append([P(f, st['label']), P('_' * 60, st['label'])])

    profile_tbl = Table(profile_data, colWidths=[55*mm, W - 55*mm])
    profile_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_CLR),
        ('BACKGROUND', (0, 0), (0, -1), SUBHEADER_BG),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(profile_tbl)
    story.append(Spacer(1, 10*mm))

    story.append(P('CARA MENGGUNAKAN BORANG', st['section']))
    cara = [
        '1. Isi identitas mahasiswa, dosen pembimbing, dan teknologi/proyek di halaman ini.',
        '2. Mulai dari TKT 1, centang setiap indikator yang terpenuhi. Isi kolom Pengukuran dengan persentase pemenuhan (0-100%).',
        '3. Hitung nilai rata-rata untuk setiap level TKT dan tulis di kolom Indikator baris Nilai Rata-rata.',
    ]
    if is_engineering:
        cara.append('4. Jika rata-rata >= 80%, lanjutkan ke level TKT berikutnya. Jika < 80%, pengukuran BERHENTI.')
    else:
        cara.append('4. Pengukuran BERHENTI pada level pertama yang rata-ratanya < 80%.')
    cara += [
        '5. TKT yang dicapai = level TKT tertinggi yang rata-ratanya >= 80%.',
        '6. Di halaman terakhir, isi ringkasan hasil dan minta tanda tangan pembimbing serta penguji (saat sidang akhir).',
    ]
    for inst in cara:
        story.append(P(inst, st['note']))
        story.append(Spacer(1, 1*mm))

    story.append(PageBreak())

    # ── Indicator tables ────────────────────────────────
    story.append(P(title, st['title']))
    story.append(P(subtitle, st['subtitle']))
    story.append(Spacer(1, 2*mm))

    for lev_data in data:
        lvl  = lev_data['level']
        inds = lev_data['indicators']
        stop = lev_data['stop_text']

        tbl_data = [[
            P(f'TKT {lvl}', st['hdr']), P('No', st['hdr_c']),
            P('Indikator', st['hdr']), P('Pengukuran<br/>0-100%', st['hdr_c']),
        ]]
        for ind in inds:
            tbl_data.append([
                '', P(str(ind['id']), st['cell_c']),
                P(ind['text'], st['cell']), P('________', st['cell_c']),
            ])
        tbl_data.append([
            '', '',
            P('<b>Nilai Rata-rata:</b> ________', st['cell']), '',
        ])

        c_w = [14*mm, 8*mm, W - 40*mm, 18*mm]
        tbl = Table(tbl_data, colWidths=c_w, repeatRows=1)

        ts = [
            ('BACKGROUND',   (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR',    (0, 0), (-1, 0), HEADER_FG),
            ('ALIGN',        (0, 0), (0, -1), 'LEFT'),
            ('ALIGN',        (1, 0), (1, -1), 'CENTER'),
            ('ALIGN',        (3, 0), (3, -1), 'RIGHT'),
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',         (0, 0), (-1, -1), 0.4, BORDER_CLR),
            ('TOPPADDING',   (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 2),
            ('LEFTPADDING',  (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('BACKGROUND',   (0, 0), (0, 0), SUBHEADER_BG),
            ('TEXTCOLOR',    (0, 0), (0, 0), HEADER_BG),
            ('FONTNAME',     (0, 0), (0, 0), FB),
            ('BACKGROUND',   (0, -1), (-1, -1), LIGHT_GRAY),
            ('FONTNAME',     (0, -1), (-1, -1), FB),
            ('SPAN',         (0, -1), (2, -1)),  # Merge No+Indikator for avg row
        ]
        for i in range(2, len(tbl_data), 2):
            ts.append(('BACKGROUND', (1, i), (-1, i), ALT_ROW_BG))

        tbl.setStyle(TableStyle(ts))
        story.append(KeepTogether([tbl]))

        # Stop / Continue note
        if stop:
            if 'BERHENTI' in stop.upper():
                ns = ParagraphStyle('stop', fontName=FB, fontSize=8, textColor=RED, spaceAfter=4, spaceBefore=2)
                story.append(P(f'X {stop}', ns))
            else:
                ns = ParagraphStyle('go', fontName=FB, fontSize=8, textColor=GREEN, spaceAfter=4, spaceBefore=2)
                story.append(P(f'OK {stop}', ns))
        else:
            story.append(Spacer(1, 3*mm))

    story.append(PageBreak())

    # ── Summary page ────────────────────────────────────
    story.append(P('RINGKASAN HASIL PENGUKURAN TKT', st['title']))
    story.append(P(subtitle, st['subtitle']))
    story.append(Spacer(1, 6*mm))

    sum_fields = [
        'Nama Teknologi / Judul', 'Bidang Teknologi',
        'Pimpinan Program / Dosen Pembimbing', 'Program Studi', 'Tanggal Pengukuran',
    ]
    sum_data = []
    for f in sum_fields:
        sum_data.append([P(f, st['label']), P('_' * 60, st['label'])])

    sum_tbl = Table(sum_data, colWidths=[55*mm, W - 55*mm])
    sum_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_CLR),
        ('BACKGROUND', (0, 0), (0, -1), SUBHEADER_BG),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 8*mm))

    result_data = [
        [P('Level TKT yang dicapai', st['label']), P('Level ________  (dari 9 level)', st['label'])],
        [P('% Komplit Indikator', st['label']),     P('________ %', st['label'])],
    ]
    res_tbl = Table(result_data, colWidths=[55*mm, W - 55*mm])
    res_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_CLR),
        ('BACKGROUND', (0, 0), (0, -1), SUBHEADER_BG),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(res_tbl)
    story.append(Spacer(1, 8*mm))

    story.append(P('TKT-METER', st['section']))
    story.append(Spacer(1, 2*mm))

    meter_hdr = [[P('Level', st['hdr_c']), P('TKT', st['hdr_c']),
                   P('Skor Rata-rata', st['hdr_c']), P('Status', st['hdr_c'])]]
    meter_rows = meter_hdr[:]
    for lev_data in data:
        lvl = lev_data['level']
        meter_rows.append([
            P(str(lvl), st['cell_c']), P(f'TKT {lvl}', st['cell_c']),
            P('________', st['cell_c']), P('[ ]', st['cell_c']),
        ])
    meter_tbl = Table(meter_rows, colWidths=[20*mm, 30*mm, 50*mm, W - 100*mm])
    meter_tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR',    (0, 0), (-1, 0), HEADER_FG),
        ('GRID',         (0, 0), (-1, -1), 0.4, BORDER_CLR),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',        (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING',   (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 2),
    ]))
    for i in range(2, len(meter_rows), 2):
        meter_tbl.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), ALT_ROW_BG)]))
    story.append(meter_tbl)
    story.append(Spacer(1, 4*mm))
    story.append(P(
        '<i>TKT yang dicapai = TKT tertinggi yang indikatornya terpenuhi (rata-rata >= 80% set point).</i>',
        st['small'],
    ))
    story.append(Spacer(1, 10*mm))

    sig_data = [
        [P('Mengetahui,', st['small']), P('', st['small']),
         P('', st['small']), P('', st['small'])],
        [P('Dosen Pembimbing', st['small']), P('Dosen Penguji 1', st['small']),
         P('Dosen Penguji 2', st['small']), P('Nama Mahasiswa', st['small'])],
        [P('<br/><br/><br/>________________________<br/>NIP.', st['small']),
         P('<br/><br/><br/>________________________<br/>NIP.', st['small']),
         P('<br/><br/><br/>________________________<br/>NIP.', st['small']),
         P('<br/><br/><br/>________________________<br/>NIM.', st['small'])],
    ]
    sig_tbl = Table(sig_data, colWidths=[W/4]*4)

    doc.build(story)
    print(f'  -> {os.path.basename(pdf_path)} ({os.path.getsize(pdf_path)} bytes)')

if __name__ == '__main__':
    os.makedirs(FORMULIR_DIR, exist_ok=True)

    build_pdf(
        xlsx_path='/Users/martinmanullang/Downloads/5. PENGUKURAN TKT SOFTWARE.xlsx',
        pdf_path  =os.path.join(FORMULIR_DIR, 'Borang TKT Software.pdf'),
        title     ='BORANG PENGUKURAN TINGKAT KESIAPAN TEKNOLOGI (TKT)',
        subtitle  ='Bidang: Perangkat Lunak / Software',
        is_engineering=False,
    )
    build_pdf(
        xlsx_path='/Users/martinmanullang/Downloads/7. PENGUKURAN TKT UMUM DAN ENGINEERING.xlsx',
        pdf_path  =os.path.join(FORMULIR_DIR, 'Borang TKT Umum dan Engineering.pdf'),
        title     ='BORANG PENGUKURAN TINGKAT KESIAPAN TEKNOLOGI (TKT)',
        subtitle  ='Bidang: Umum & Hard Engineering',
        is_engineering=True,
    )
